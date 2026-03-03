"""
Import external nutrition databases (CIQUAL 2025, MEXT Japan 7th) into index files
compatible with NutritionMatcher.

Usage:
    python import_nutrition_sources.py --ciqual-xml PATH --ciqual-xlsx PATH --mext-xlsx PATH

Source files (not included in repo, download manually):
    CIQUAL 2025: https://entrepot.recherche.data.gouv.fr/dataset.xhtml?persistentId=doi:10.57745/RDMHWY
        - alim_2025.xml  (food names FR+EN)
        - Table Ciqual 2025_FR_2025_11_03.xlsx  (nutrition values)
    MEXT Japan 7th: https://www.mext.go.jp/en/.../1374049_1r12_1.xlsx
"""

import argparse
import json
import logging
import re
import xml.etree.ElementTree as ET
from pathlib import Path
from typing import Any, Optional

import openpyxl

logger = logging.getLogger(__name__)

OUTPUT_DIR = Path(__file__).parent.parent / "packages" / "recipe_scraper" / "src" / "recipe_scraper" / "data"


def _parse_ciqual_value(raw: Any) -> Optional[float]:
    """Parse a CIQUAL nutrition cell value (string) into a float or None."""
    if raw is None:
        return None
    s = str(raw).strip()
    if s in ("-", "", "N"):
        return None
    if s.lower() == "traces":
        return 0.0
    m = re.match(r"<\s*([\d.,]+)", s)
    if m:
        return float(m.group(1).replace(",", ".")) / 2
    try:
        return float(s.replace(",", "."))
    except ValueError:
        logger.warning(f"CIQUAL: unparseable value '{s}'")
        return None


def _parse_mext_value(raw: Any) -> Optional[float]:
    """Parse a MEXT nutrition cell value into a float or None."""
    if raw is None:
        return None
    s = str(raw).strip()
    if s in ("-", "", "Tr", "—", "(0)", "0"):
        if s == "(0)":
            return 0.0
        if s == "0":
            return 0.0
        if s == "Tr":
            return 0.0
        return None
    s = s.strip("()")
    try:
        return float(s)
    except ValueError:
        logger.warning(f"MEXT: unparseable value '{s}'")
        return None


def _normalize_mext_name(raw_name: str) -> tuple[str, list[str]]:
    """
    Normalize a verbose MEXT food name into a clean primary name + alt names.

    Extracts quoted Japanese terms and short parentheticals as aliases.
    """
    name = raw_name
    name = re.sub(r"[\[［][^\]］]*[\]］]", "", name)
    name = name.replace("\n", " ")

    quoted = re.findall(r'"([^"]+)"', name)
    parens = re.findall(r"\(([^)]{3,30})\)", name)

    clean = re.sub(r'"[^"]*"', "", name)
    clean = re.sub(r"\([^)]*\)", "", clean)
    clean = re.sub(r"\s*,\s*,", ",", clean)
    clean = re.sub(r",\s*$", "", clean)
    clean = re.sub(r"^\s*,\s*", "", clean)
    clean = re.sub(r"\s+", " ", clean).strip().lower()
    clean = re.sub(r"\*", "", clean)

    alts = [q.lower().strip() for q in quoted + parens if len(q.strip()) > 2]
    return clean, alts


def convert_ciqual(xml_path: Path, xlsx_path: Path) -> list[dict]:
    """Convert CIQUAL 2025 XML (names) + Excel (nutrition) into index entries."""
    logger.info(f"Loading CIQUAL XML from {xml_path}")
    tree = ET.parse(xml_path)
    root = tree.getroot()

    names_map: dict[str, dict[str, str]] = {}
    for elem in root:
        code = (elem.find("alim_code").text or "").strip()
        nom_fr = (elem.find("alim_nom_fr").text or "").strip()
        nom_en = (elem.find("alim_nom_eng").text or "").strip()
        names_map[code] = {"name_en": nom_en.lower(), "name_fr": nom_fr.lower()}

    logger.info(f"Loaded {len(names_map)} food names from XML")

    logger.info(f"Loading CIQUAL Excel from {xlsx_path}")
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active

    raw_headers = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
    headers = [str(h or "").replace("\n", " ").strip().lower() for h in raw_headers]

    col_map = {}
    for i, h in enumerate(headers):
        if "alim_code" in h:
            col_map["code"] = i
        elif "kcal" in h and "1169" in h and "kcal" not in col_map:
            col_map["kcal"] = i
        elif "protéines" in h and "jones" in h and "protein" not in col_map:
            col_map["protein"] = i
        elif h.startswith("glucides"):
            col_map["carbs"] = i
        elif h.startswith("lipides"):
            col_map["fat"] = i
        elif h.startswith("sucres"):
            col_map["sugar"] = i
        elif h.startswith("fibres"):
            col_map["fiber"] = i
        elif h.startswith("ag saturés") or ("ag" in h and "saturés" in h and "sat_fat" not in col_map):
            col_map["sat_fat"] = i
        elif h.startswith("calcium") and "calcium_mg" not in col_map:
            col_map["calcium_mg"] = i
        elif h.startswith("fer") and "iron_mg" not in col_map:
            col_map["iron_mg"] = i
        elif h.startswith("magnésium") and "magnesium_mg" not in col_map:
            col_map["magnesium_mg"] = i
        elif h.startswith("potassium") and "potassium_mg" not in col_map:
            col_map["potassium_mg"] = i
        elif h.startswith("sodium") and "sodium_mg" not in col_map:
            col_map["sodium_mg"] = i
        elif h.startswith("zinc") and "zinc_mg" not in col_map:
            col_map["zinc_mg"] = i

    logger.info(f"CIQUAL column mapping: {col_map}")

    entries = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        code = str(row[col_map["code"]]).strip() if row[col_map["code"]] else None
        if not code or code not in names_map:
            continue

        names = names_map[code]
        if not names["name_en"]:
            continue

        entry = {
            "id": f"ciqual_{code}",
            "name": names["name_en"],
            "alt": [names["name_fr"]] if names["name_fr"] and names["name_fr"] != names["name_en"] else [],
            "kcal": _parse_ciqual_value(row[col_map.get("kcal", -1)] if "kcal" in col_map else None),
            "protein": _parse_ciqual_value(row[col_map.get("protein", -1)] if "protein" in col_map else None),
            "fat": _parse_ciqual_value(row[col_map.get("fat", -1)] if "fat" in col_map else None),
            "carbs": _parse_ciqual_value(row[col_map.get("carbs", -1)] if "carbs" in col_map else None),
            "fiber": _parse_ciqual_value(row[col_map.get("fiber", -1)] if "fiber" in col_map else None),
            "sugar": _parse_ciqual_value(row[col_map.get("sugar", -1)] if "sugar" in col_map else None),
            "sat_fat": _parse_ciqual_value(row[col_map.get("sat_fat", -1)] if "sat_fat" in col_map else None),
            "calcium_mg": _parse_ciqual_value(row[col_map.get("calcium_mg", -1)] if "calcium_mg" in col_map else None),
            "iron_mg": _parse_ciqual_value(row[col_map.get("iron_mg", -1)] if "iron_mg" in col_map else None),
            "magnesium_mg": _parse_ciqual_value(row[col_map.get("magnesium_mg", -1)] if "magnesium_mg" in col_map else None),
            "potassium_mg": _parse_ciqual_value(row[col_map.get("potassium_mg", -1)] if "potassium_mg" in col_map else None),
            "sodium_mg": _parse_ciqual_value(row[col_map.get("sodium_mg", -1)] if "sodium_mg" in col_map else None),
            "zinc_mg": _parse_ciqual_value(row[col_map.get("zinc_mg", -1)] if "zinc_mg" in col_map else None),
            "source": "ciqual",
        }

        if entry["kcal"] is not None:
            entries.append(entry)

    logger.info(f"Converted {len(entries)} CIQUAL entries (with kcal)")
    return entries


def convert_mext(xlsx_path: Path) -> list[dict]:
    """Convert MEXT Japan 7th edition Excel into index entries."""
    logger.info(f"Loading MEXT Excel from {xlsx_path}")
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb["Table"]

    header_row = list(ws.iter_rows(min_row=6, max_row=6, values_only=True))[0]
    headers = [str(h or "").strip() for h in header_row]

    col_map = {}
    for i, h in enumerate(headers):
        hl = h.lower()
        if "food and description" in hl:
            col_map["name"] = i
        elif h == "Energy (kcal)":
            col_map["kcal"] = i
        elif "protein, calculated from" in hl and "reference" in hl:
            col_map["protein"] = i
        elif hl.startswith("lipid"):
            col_map["fat"] = i
        elif "carbohydrate, total" in hl:
            col_map["carbs"] = i
        elif "dietary fiber" in hl:
            col_map["fiber"] = i
        elif "fatty acid, saturated" in hl or h == "Fatty acid, saturated":
            col_map["sat_fat"] = i
        elif hl == "calcium" or hl.startswith("calcium"):
            if "calcium_mg" not in col_map:
                col_map["calcium_mg"] = i
        elif hl == "iron" or hl.startswith("iron"):
            if "iron_mg" not in col_map:
                col_map["iron_mg"] = i
        elif hl == "magnesium" or hl.startswith("magnesium"):
            if "magnesium_mg" not in col_map:
                col_map["magnesium_mg"] = i
        elif hl == "potassium" or hl.startswith("potassium"):
            if "potassium_mg" not in col_map:
                col_map["potassium_mg"] = i
        elif hl == "sodium" or hl.startswith("sodium"):
            if "sodium_mg" not in col_map:
                col_map["sodium_mg"] = i
        elif hl == "zinc" or hl.startswith("zinc"):
            if "zinc_mg" not in col_map:
                col_map["zinc_mg"] = i

    if "kcal" not in col_map:
        for i, h in enumerate(headers):
            if "kcal" in h.lower():
                col_map["kcal"] = i
                break

    logger.info(f"MEXT column mapping: {col_map}")

    item_no_col = None
    for i, h in enumerate(headers):
        if "item" in h.lower() and "no" in h.lower():
            item_no_col = i
            break

    entries = []
    for row in ws.iter_rows(min_row=9, values_only=True):
        raw_name = row[col_map.get("name", 3)]
        if not raw_name or not str(raw_name).strip() or str(raw_name).strip() == "-":
            continue

        raw_name = str(raw_name).strip()
        name, alts = _normalize_mext_name(raw_name)

        if not name:
            continue

        item_no = str(row[item_no_col]).strip() if item_no_col is not None and row[item_no_col] else ""

        entry = {
            "id": f"mext_{item_no}" if item_no else f"mext_{name[:30].replace(' ', '_')}",
            "name": name,
            "alt": alts,
            "kcal": _parse_mext_value(row[col_map.get("kcal", -1)] if "kcal" in col_map else None),
            "protein": _parse_mext_value(row[col_map.get("protein", -1)] if "protein" in col_map else None),
            "fat": _parse_mext_value(row[col_map.get("fat", -1)] if "fat" in col_map else None),
            "carbs": _parse_mext_value(row[col_map.get("carbs", -1)] if "carbs" in col_map else None),
            "fiber": _parse_mext_value(row[col_map.get("fiber", -1)] if "fiber" in col_map else None),
            "sugar": None,
            "sat_fat": _parse_mext_value(row[col_map.get("sat_fat", -1)] if "sat_fat" in col_map else None),
            "calcium_mg": _parse_mext_value(row[col_map.get("calcium_mg", -1)] if "calcium_mg" in col_map else None),
            "iron_mg": _parse_mext_value(row[col_map.get("iron_mg", -1)] if "iron_mg" in col_map else None),
            "magnesium_mg": _parse_mext_value(row[col_map.get("magnesium_mg", -1)] if "magnesium_mg" in col_map else None),
            "potassium_mg": _parse_mext_value(row[col_map.get("potassium_mg", -1)] if "potassium_mg" in col_map else None),
            "sodium_mg": _parse_mext_value(row[col_map.get("sodium_mg", -1)] if "sodium_mg" in col_map else None),
            "zinc_mg": _parse_mext_value(row[col_map.get("zinc_mg", -1)] if "zinc_mg" in col_map else None),
            "source": "mext",
        }

        if entry["kcal"] is not None:
            entries.append(entry)

    logger.info(f"Converted {len(entries)} MEXT entries (with kcal)")
    return entries


def main():
    parser = argparse.ArgumentParser(description="Import external nutrition databases")
    parser.add_argument("--ciqual-xml", type=Path, help="Path to CIQUAL alim_2025.xml")
    parser.add_argument("--ciqual-xlsx", type=Path, help="Path to CIQUAL 2025 Excel file")
    parser.add_argument("--mext-xlsx", type=Path, help="Path to MEXT Japan 7th ed Excel file")
    parser.add_argument("--output-dir", type=Path, default=OUTPUT_DIR, help="Output directory")
    args = parser.parse_args()

    logging.basicConfig(level=logging.INFO, format="%(levelname)s: %(message)s")

    args.output_dir.mkdir(parents=True, exist_ok=True)

    if args.ciqual_xml and args.ciqual_xlsx:
        entries = convert_ciqual(args.ciqual_xml, args.ciqual_xlsx)
        out = args.output_dir / "ciqual_index.json"
        with open(out, "w", encoding="utf-8") as f:
            json.dump(entries, f, ensure_ascii=False)
        logger.info(f"Wrote {len(entries)} entries to {out}")
    else:
        logger.info("Skipping CIQUAL (no --ciqual-xml / --ciqual-xlsx)")

    if args.mext_xlsx:
        entries = convert_mext(args.mext_xlsx)
        out = args.output_dir / "mext_index.json"
        with open(out, "w", encoding="utf-8") as f:
            json.dump(entries, f, ensure_ascii=False)
        logger.info(f"Wrote {len(entries)} entries to {out}")
    else:
        logger.info("Skipping MEXT (no --mext-xlsx)")


if __name__ == "__main__":
    main()
